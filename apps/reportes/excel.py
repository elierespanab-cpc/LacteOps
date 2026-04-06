from datetime import datetime

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def exportar_excel(titulo, columnas, filas, filename=None, empresa=None, parametros=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center")

    row = 1
    nombre_empresa = "LacteOps"
    if empresa:
        nombre_empresa = (
            getattr(empresa, "nombre_empresa", None)
            or getattr(empresa, "nombre", None)
            or "LacteOps"
        )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columnas))
    c = ws.cell(row=row, column=1, value=nombre_empresa)
    c.font = Font(bold=True, size=14, color="1F3864")
    row += 1

    if empresa:
        rif = getattr(empresa, "rif", "")
        direccion = getattr(empresa, "direccion", "")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columnas))
        ws.cell(row=row, column=1, value=f"RIF: {rif} | Dirección: {direccion}").font = Font(size=10)
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columnas))
    ws.cell(row=row, column=1, value=titulo).font = Font(bold=True, size=12)
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columnas))
    ws.cell(
        row=row,
        column=1,
        value=f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
    ).font = Font(size=10)
    row += 1

    if parametros:
        for clave, valor in parametros.items():
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columnas))
            cell = ws.cell(row=row, column=1, value=f"{clave}: {valor}")
            cell.font = Font(italic=True, size=10)
            row += 1

    row += 1

    for ci, col in enumerate(columnas, 1):
        c = ws.cell(row=row, column=ci, value=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
    row += 1

    for fila in filas:
        for ci, val in enumerate(fila, 1):
            ws.cell(row=row, column=ci, value=val)
        row += 1

    for idx in range(1, len(columnas) + 1):
        max_len = len(str(columnas[idx - 1]))
        for row_cells in ws.iter_rows(min_col=idx, max_col=idx):
            cell = row_cells[0]
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 14), 40)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    fn = filename or f"{titulo}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{fn}"'
    wb.save(response)
    return response
